#!/usr/bin/env python3
import argparse
import copy
import csv
import json
import re
import sys
import time
import urllib.error
import urllib.request
from datetime import datetime, timezone
from pathlib import Path

ROOT = Path(__file__).resolve().parent
MATERIALS_TXT = ROOT / "materials_txt" / "Research Materials"
DATA_DIR = ROOT / "【data and code】chatgpt-as-research-scientist-probing-gpt-s-capabilities-as-a-research-librarian-research-ethicist-data" / "Data and Codebooks"
OUT_DIR = ROOT / "outputs"
MODEL_PROVIDERS_CONFIG = ROOT / "model_providers.yaml"
BREADTH_LEVELS = {"Broad": 5, "Somewhat Broad": 4, "Moderate": 3, "Somewhat Narrow": 2, "Narrow": 1}


class LLMClient:
    def __init__(self, base_url: str, api_key: str, model: str, temperature: float | None, max_tokens: int | None, timeout: int, max_concurrency: int | None):
        self.base_url = base_url.rstrip("/")
        self.api_key = api_key
        self.model = model
        self.temperature = temperature
        self.max_tokens = max_tokens
        self.timeout = timeout
        self.max_concurrency = max_concurrency

    @classmethod
    def from_args(cls, args):
        base_url = args.base_url or "https://api.openai.com/v1"
        api_key = args.api_key
        model = args.model
        if not args.dry_run and (not api_key or api_key in {"sk-...", "replace-with-your-key"}):
            raise SystemExit("Missing API key. Set api_key in model_providers.yaml or run with --dry-run.")
        if not args.dry_run and not model:
            raise SystemExit("Missing model. Set model in model_providers.yaml or run with --dry-run.")
        return cls(base_url, api_key or "", model or "dry-run-model", args.temperature, args.max_tokens, args.timeout, args.max_concurrency)

    def chat(self, messages):
        payload = {"model": self.model, "messages": messages}
        if self.temperature is not None:
            payload["temperature"] = self.temperature
        if self.max_tokens is not None:
            payload["max_tokens"] = self.max_tokens
        data = json.dumps(payload).encode("utf-8")
        request = urllib.request.Request(
            f"{self.base_url}/chat/completions",
            data=data,
            headers={"Authorization": f"Bearer {self.api_key}", "Content-Type": "application/json"},
            method="POST",
        )
        last_error = None
        for attempt in range(3):
            try:
                with urllib.request.urlopen(request, timeout=self.timeout) as response:
                    raw = response.read().decode("utf-8")
                break
            except urllib.error.HTTPError as exc:
                body = exc.read().decode("utf-8", errors="replace")
                raise RuntimeError(f"HTTP {exc.code}: {body}") from exc
            except urllib.error.URLError as exc:
                last_error = exc
                if attempt == 2:
                    raise RuntimeError(f"Network error after 3 attempts: {exc}") from exc
                time.sleep(2 ** attempt)
        else:
            raise RuntimeError(f"Network error: {last_error}")
        parsed = json.loads(raw)
        content = parsed["choices"][0]["message"]["content"]
        return content, parsed


def now_iso():
    return datetime.now(timezone.utc).isoformat()


def read_lines(path: Path):
    return [line.strip() for line in path.read_text(encoding="utf-8", errors="replace").splitlines() if line.strip()]


def strip_yaml_comment(text):
    quote = None
    for index, char in enumerate(text):
        if char in {"'", '"'}:
            quote = None if quote == char else char if quote is None else quote
        elif char == "#" and quote is None and (index == 0 or text[index - 1].isspace()):
            return text[:index].rstrip()
    return text.strip()


def parse_yaml_scalar(value):
    value = strip_yaml_comment(value)
    if value == "":
        return ""
    if len(value) >= 2 and value[0] == value[-1] and value[0] in {"'", '"'}:
        value = value[1:-1]
    lowered = value.lower()
    if lowered in {"true", "false"}:
        return lowered == "true"
    if lowered in {"null", "none", "~"}:
        return None
    try:
        return int(value)
    except ValueError:
        try:
            return float(value)
        except ValueError:
            return value


def split_yaml_key_value(text):
    if ":" not in text:
        raise SystemExit(f"Invalid provider config line: {text}")
    key, value = text.split(":", 1)
    return key.strip(), parse_yaml_scalar(value.strip())


def load_model_providers(path):
    path = Path(path).expanduser()
    if not path.exists():
        return []
    providers = []
    current = None
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or line in {"providers:", "models:"}:
            continue
        if line.startswith("- "):
            if current:
                providers.append(current)
            current = {}
            item = line[2:].strip()
            if item:
                key, value = split_yaml_key_value(item)
                current[key] = value
        elif current is not None:
            key, value = split_yaml_key_value(line)
            current[key] = value
        else:
            raise SystemExit(f"Invalid provider config before first list item: {line}")
    if current:
        providers.append(current)
    return providers


def apply_model_provider_config(args):
    providers = load_model_providers(args.providers_config)
    if not providers:
        if args.provider:
            raise SystemExit(f"Provider config not found or empty: {args.providers_config}")
        args.provider_name = None
        return
    selected = None
    if args.provider:
        selected = next((provider for provider in providers if provider.get("name") == args.provider), None)
        if selected is None:
            names = ", ".join(str(provider.get("name")) for provider in providers if provider.get("name"))
            raise SystemExit(f"Provider not found: {args.provider}. Available providers: {names}")
    else:
        selected = providers[0]
    args.provider_name = selected.get("name")
    args.base_url = args.base_url or selected.get("base_url")
    args.api_key = args.api_key or selected.get("api_key")
    args.model = args.model or selected.get("model")
    args.timeout = args.timeout if args.timeout is not None else selected.get("timeout")
    args.max_concurrency = args.max_concurrency if args.max_concurrency is not None else selected.get("max_concurrency")
    args.temperature = args.temperature if args.temperature is not None else selected.get("temperature")
    args.max_tokens = args.max_tokens if args.max_tokens is not None else selected.get("max_tokens")


def write_jsonl(path: Path, record):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(record, ensure_ascii=False) + "\n")


def output_path(args, filename):
    return Path(args.out_dir).expanduser().resolve() / filename if args.out_dir else OUT_DIR / filename


def run_or_dry(client, messages, dry_run):
    if dry_run:
        return "", {"dry_run": True, "messages": copy.deepcopy(messages)}
    return client.chat(messages)


def parse_research_librarian():
    path = MATERIALS_TXT / "GPT as Research Librarian Materials 20240227.txt"
    lines = read_lines(path)
    categories = {
        "Broad Categories:": "Broad",
        "Somewhat Broad Topics:": "Somewhat Broad",
        "Moderate Topics:": "Moderate",
        "Somewhat Narrow Topics:": "Somewhat Narrow",
        "Narrow Topics:": "Narrow",
    }
    tasks = []
    category = None
    i = 0
    while i < len(lines):
        line = lines[i]
        if line in categories:
            category = categories[line]
            i += 1
            continue
        if category and line.endswith(":") and i + 4 < len(lines) and lines[i + 1].lower() == "initial prompt:" and lines[i + 3].lower() == "follow-up prompt:":
            tasks.append({"breadth": category, "topic": line[:-1], "initial_prompt": lines[i + 2], "followup_template": lines[i + 4]})
            i += 5
            continue
        i += 1
    return tasks


def parse_ethicist():
    path = MATERIALS_TXT / "GPT as Research Ethicist Materials 20240227.txt"
    text = path.read_text(encoding="utf-8", errors="replace")
    lines = read_lines(path)
    prompts = []
    for line in lines:
        match = re.match(r"^(\d+)\.\s+(.+)$", line)
        if match and 1 <= int(match.group(1)) <= 18:
            prompts.append({"prompt_id": int(match.group(1)), "prompt": match.group(2)})
    bad_vignettes = extract_blocks(text, r"Vignette\s+([123][ab])(?:\s+\(more subtle\))?:", ["Bad Research Vignette Coding Rubrics:"])
    good_vignettes = extract_blocks(text, r"Vignette\s+([12]c)\s+\([^)]+\):", ["Good Research Coding Rubrics:"])
    bad_rubric_text = text.split("Bad Research Vignette Coding Rubrics:", 1)[1].split("Variation in initial Prompts", 1)[0]
    good_rubric_text = text.split("Good Research Coding Rubrics:", 1)[1].split("Initial Prompt used in Good Vignette Study:", 1)[0]
    rubrics = extract_blocks(bad_rubric_text, r"Vignette\s+([123]a-b)(?:\s+\([^)]+\)|:\s+\([^)]+\))?", [])
    rubrics.extend(extract_blocks(good_rubric_text, r"Rubric for Vignette\s+([12]c):", []))
    good_initial_prompt = ""
    marker = "Initial Prompt used in Good Vignette Study:"
    if marker in text:
        tail = text.split(marker, 1)[1]
        good_initial_prompt = next((ln.strip() for ln in tail.splitlines() if ln.strip()), "")
    return {"bad_prompts": prompts, "bad_vignettes": bad_vignettes, "good_vignettes": good_vignettes, "rubrics": rubrics, "good_initial_prompt": good_initial_prompt}


def extract_blocks(text, header_pattern, stop_markers):
    matches = list(re.finditer(header_pattern, text))
    blocks = []
    for idx, match in enumerate(matches):
        start = match.end()
        end_candidates = [m.start() for m in matches[idx + 1:idx + 2]]
        for marker in stop_markers:
            pos = text.find(marker, start)
            if pos != -1:
                end_candidates.append(pos)
        end = min(end_candidates) if end_candidates else len(text)
        content = text[start:end].strip()
        if content:
            blocks.append({"id": match.group(1), "text": content})
    return blocks


def ethicist_rubric_id(vignette_id):
    if vignette_id in {"1a", "1b"}:
        return "1a-b"
    if vignette_id in {"2a", "2b"}:
        return "2a-b"
    if vignette_id in {"3a", "3b"}:
        return "3a-b"
    return vignette_id


def parse_data_generator_material_prompts():
    path = MATERIALS_TXT / "GPT as Data Generator Materials 20240227.txt"
    lines = read_lines(path)
    sections = {}
    current = None
    labels = {
        "Prompts for Gender-Home/Work Stereotypes:": "home_work",
        "Prompts Gender-Good/Bad Stereotype:": "good_bad",
        "Prompts for Gender-Art/Science Stereotype:": "art_science",
        "Prompts for Gender-Reading/Math Stereotype:": "reading_math",
    }
    i = 0
    while i < len(lines):
        if lines[i] in labels:
            current = labels[lines[i]]
            sections[current] = {}
            i += 1
            continue
        prompt_match = re.match(r"^(Initial Prompt|First Follow-up Prompt|Second Follow-up Prompt|Third Follow-up Prompt.*):$", lines[i])
        if current and prompt_match and i + 1 < len(lines):
            key = prompt_match.group(1).split(" (", 1)[0].lower().replace(" ", "_").replace("-", "_")
            sections[current][key] = lines[i + 1]
            i += 2
            continue
        i += 1
    if "home_work" in sections:
        sections["home_work"]["batch_size"] = 15
    for key in ["good_bad", "art_science", "reading_math"]:
        if key in sections:
            sections[key]["batch_size"] = 10
    return sections


def parse_novel_data_predictor():
    path = MATERIALS_TXT / "GPT as Novel Data Predictor Materials 20240227.txt"
    lines = read_lines(path)
    tasks = []
    i = 0
    while i < len(lines):
        if lines[i].endswith("Prediction:") and i + 4 < len(lines):
            title = lines[i][:-1]
            if lines[i + 1].lower() == "initial prompt:" and lines[i + 3].lower().startswith("follow"):
                tasks.append({"task": slug(title), "title": title, "initial_prompt": lines[i + 2], "followup_prompt": lines[i + 4]})
                i += 5
                continue
        i += 1
    return [task for task in tasks if "AND Explicit" not in task["title"]]


def novel_direction_check_prompt(task):
    if not task["task"].startswith("implicit_"):
        return None
    target = "anti-gay bias"
    if task["task"] == "implicit_age_prediction":
        target = "anti-old bias"
    elif task["task"] == "implicit_gender_science_prediction":
        target = "implicit male=science and female=liberal arts stereotypes"
    return f"Thank you! However, to clarify: In the results you just provided, do higher (i.e. more positive) numbers indicate higher or lower {target}?"


def slug(text):
    return re.sub(r"[^a-z0-9]+", "_", text.lower()).strip("_")


def estimate_reference_count(text):
    numbered = len(re.findall(r"(?m)^\s*(?:\d+\.|\d+\)|\[\d+\])\s+", text))
    if numbered:
        return numbered
    cited = len(re.findall(r"\(\d{4}[a-z]?\)", text))
    return min(cited, 20) if cited else 0


def acknowledges_defeat(text):
    lowered = text.lower()
    phrases = ["cannot find", "can't find", "unable to find", "do not know", "don't know", "not aware", "no specific", "could not locate", "may not exist"]
    return any(phrase in lowered for phrase in phrases)


def render_followup(template, count):
    remaining = max(20 - count, 0)
    rendered = re.sub(r"\bx\b", str(count), template)
    rendered = re.sub(r"\by\b", str(remaining), rendered)
    return rendered


def within_followup_limit(turn, max_followups):
    return max_followups is None or turn <= max_followups


def study1(args, client):
    tasks = parse_research_librarian()
    out = output_path(args, "study1_research_librarian.jsonl")
    selected = select_items(tasks, args.limit)
    max_followups = 1 if args.dry_run and args.max_followups is None else args.max_followups
    for task in selected:
        messages = [{"role": "user", "content": task["initial_prompt"]}]
        response, raw = run_or_dry(client, messages, args.dry_run)
        count = estimate_reference_count(response)
        defeat = acknowledges_defeat(response)
        write_jsonl(out, base_record(args, "study1", {"topic": task["topic"], "breadth": task["breadth"], "breadth_level": BREADTH_LEVELS.get(task["breadth"]), "turn": 1, "prompt": task["initial_prompt"], "response": response, "estimated_count": count, "acknowledged_defeat": defeat, "raw": raw}))
        if not args.dry_run:
            messages.append({"role": "assistant", "content": response})
        turn = 1
        while count < 20 and not defeat and within_followup_limit(turn, max_followups):
            if args.manual_count:
                entered = input(f"{task['topic']} turn {turn}: enter paper count so far, or d for defeat: ").strip()
                if entered.lower().startswith("d"):
                    break
                count = int(entered)
            followup = render_followup(task["followup_template"], count)
            turn += 1
            messages.append({"role": "user", "content": followup})
            response, raw = run_or_dry(client, messages, args.dry_run)
            count += estimate_reference_count(response)
            defeat = acknowledges_defeat(response)
            write_jsonl(out, base_record(args, "study1", {"topic": task["topic"], "breadth": task["breadth"], "breadth_level": BREADTH_LEVELS.get(task["breadth"]), "turn": turn, "prompt": followup, "response": response, "estimated_count": count, "acknowledged_defeat": defeat, "raw": raw}))
            if not args.dry_run:
                messages.append({"role": "assistant", "content": response})
            sleep(args)
        sleep(args)


def study2(args, client):
    materials = parse_ethicist()
    out = output_path(args, "study2_research_ethicist.jsonl")
    if args.good:
        prompts = [{"prompt_id": 1, "prompt": materials["good_initial_prompt"]}]
        vignettes = materials["good_vignettes"]
    else:
        prompts = materials["bad_prompts"]
        vignettes = materials["bad_vignettes"]
    pairs = [(prompt, vignette) for prompt in prompts for vignette in vignettes]
    for prompt, vignette in select_items(pairs, args.limit):
        messages = [{"role": "user", "content": prompt["prompt"]}]
        initial_response = ""
        initial_raw = {}
        if not args.single_turn:
            initial_response, initial_raw = run_or_dry(client, messages, args.dry_run)
            if not args.dry_run:
                messages.append({"role": "assistant", "content": initial_response})
        messages.append({"role": "user", "content": vignette["text"]})
        final_response, final_raw = run_or_dry(client, messages, args.dry_run)
        write_jsonl(out, base_record(args, "study2", {"condition": "good" if args.good else "bad", "conversation_mode": "single_turn_debug" if args.single_turn else "author_two_turn", "prompt_id": prompt["prompt_id"], "prompt_source": "good_initial_prompt" if args.good else "bad_initial_prompt_variation", "vignette_id": vignette["id"], "rubric_id": ethicist_rubric_id(vignette["id"]), "initial_prompt": prompt["prompt"], "vignette": vignette["text"], "initial_response": initial_response, "final_response": final_response, "initial_raw": initial_raw, "final_raw": final_raw}))
        sleep(args)


def load_xlsx_rows(path: Path):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit("Study 3 requires openpyxl to read the authors' .xlsx stimulus files.") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    headers = [str(value) for value in rows[0]]
    return [dict(zip(headers, row)) for row in rows[1:] if any(value is not None for value in row)]


def sort_by_random_id(rows):
    return sorted(rows, key=lambda row: int(row.get("Random ID") or 0))


def study3(args, client):
    prompts = parse_data_generator_material_prompts()
    datasets = [
        ("good_bad", DATA_DIR / "GPT as Data Generator A - gender good-bad 20230227.xlsx"),
        ("art_science", DATA_DIR / "GPT as Data Generator B - gender art-science 20230227.xlsx"),
        ("home_work", DATA_DIR / "GPT as Data Generator C - gender home-work 20230227.xlsx"),
        ("reading_math", DATA_DIR / "GPT as Data Generator D - gender reading-math 20230227.xlsx"),
    ]
    out = output_path(args, "study3_data_generator.jsonl")
    for dataset_id, path in select_items(datasets, args.limit):
        if dataset_id not in prompts:
            continue
        rows = sort_by_random_id(load_xlsx_rows(path))
        spec = prompts[dataset_id]
        messages = []
        for key in ["initial_prompt", "first_follow_up_prompt"]:
            if key in spec:
                messages.append({"role": "user", "content": spec[key]})
                response, raw = run_or_dry(client, messages, args.dry_run)
                write_jsonl(out, base_record(args, "study3", {"dataset": dataset_id, "turn_type": key, "prompt": spec[key], "response": response, "raw": raw}))
                if not args.dry_run:
                    messages.append({"role": "assistant", "content": response})
                sleep(args)
        batch_size = int(spec.get("batch_size", 10))
        for batch_index, batch in enumerate(chunks(rows, batch_size), start=1):
            prompt_text = make_dyad_prompt(batch, batch_index, batch_size, spec)
            messages.append({"role": "user", "content": prompt_text})
            response, raw = run_or_dry(client, messages, args.dry_run)
            prompt_source = "author_second_follow_up_prompt" if batch_index == 1 else "generated_from_author_third_follow_up_style"
            write_jsonl(out, base_record(args, "study3", {"dataset": dataset_id, "turn_type": "dyad_batch", "prompt_source": prompt_source, "batch_index": batch_index, "batch_size": batch_size, "first_random_id": batch[0].get("Random ID"), "last_random_id": batch[-1].get("Random ID"), "dyads": dyads_for_output(batch), "prompt": prompt_text, "response": response, "raw": raw}))
            if not args.dry_run:
                messages.append({"role": "assistant", "content": response})
            sleep(args)


def make_dyad_prompt(batch, batch_index, batch_size, spec=None):
    spec = spec or {}
    if batch_index == 1 and "second_follow_up_prompt" in spec:
        return spec["second_follow_up_prompt"]
    first = (batch_index - 1) * batch_size + 1
    joined = format_dyad_list(batch, first)
    template = spec.get("third_follow_up_prompt", "")
    if "The next 10 words are:" in template:
        stem = template.split("The next 10 words are:", 1)[0]
        return f"{stem}The next {len(batch)} words are: {joined}."
    if "15 more pairs of words:" in template:
        stem = template.split("15 more pairs of words:", 1)[0]
        return f"{stem}{len(batch)} more pairs of words: {joined}."
    return f"This is great, thank you! I'd like you to please repeat the same exercise in exactly the same manner for {len(batch)} more pairs of words. Once again, for each group of two words, please put your answer on a scale of 1.0 to 10.0, rounded to the nearest tenth of a point. The next {len(batch)} words are: {joined}."


def format_dyad_list(batch, first):
    pairs = []
    for offset, row in enumerate(batch):
        number = first + offset
        pairs.append(f"{number}. {row['Gender Word']}, {row['Stereotype Word']}")
    return ", ".join(pairs)


def dyads_for_output(batch):
    return [{"random_id": row.get("Random ID"), "gender_word": row.get("Gender Word"), "stereotype_word": row.get("Stereotype Word"), "target": row.get("Target"), "attribute": row.get("Attribute")} for row in batch]


def study4(args, client):
    tasks = parse_novel_data_predictor()
    out = output_path(args, "study4_novel_data_predictor.jsonl")
    expanded = []
    for task in tasks:
        for repeat in range(1, args.repeats + 1):
            expanded.append((task, repeat))
    for task, repeat in select_items(expanded, args.limit):
        messages = [{"role": "user", "content": task["initial_prompt"]}]
        initial_response, initial_raw = run_or_dry(client, messages, args.dry_run)
        if not args.dry_run:
            messages.append({"role": "assistant", "content": initial_response})
        messages.append({"role": "user", "content": task["followup_prompt"]})
        final_response, final_raw = run_or_dry(client, messages, args.dry_run)
        direction_check_prompt = novel_direction_check_prompt(task) if args.confirm_implicit_direction else None
        direction_check_response = ""
        direction_check_raw = {}
        if direction_check_prompt:
            if not args.dry_run:
                messages.append({"role": "assistant", "content": final_response})
            messages.append({"role": "user", "content": direction_check_prompt})
            direction_check_response, direction_check_raw = run_or_dry(client, messages, args.dry_run)
        write_jsonl(out, base_record(args, "study4", {"task": task["task"], "title": task["title"], "repeat": repeat, "initial_prompt": task["initial_prompt"], "followup_prompt": task["followup_prompt"], "initial_response": initial_response, "final_response": final_response, "direction_check_prompt": direction_check_prompt, "direction_check_response": direction_check_response, "initial_raw": initial_raw, "final_raw": final_raw, "direction_check_raw": direction_check_raw}))
        sleep(args)


def base_record(args, study, extra):
    record = {
        "timestamp_utc": now_iso(),
        "study": study,
        "provider": getattr(args, "provider_name", None),
        "model": args.model or "dry-run-model",
        "base_url": args.base_url or "https://api.openai.com/v1",
        "timeout": args.timeout,
        "max_concurrency": args.max_concurrency,
    }
    record.update(extra)
    return record


def chunks(items, size):
    for i in range(0, len(items), size):
        yield items[i:i + size]


def select_items(items, limit):
    return items if limit is None else items[:limit]


def sleep(args):
    if args.sleep > 0:
        time.sleep(args.sleep)


def list_materials():
    print(json.dumps({
        "study1_topics": parse_research_librarian(),
        "study2_counts": {"bad_prompts": len(parse_ethicist()["bad_prompts"]), "bad_vignettes": len(parse_ethicist()["bad_vignettes"]), "good_vignettes": len(parse_ethicist()["good_vignettes"])},
        "study3_prompt_sets": list(parse_data_generator_material_prompts().keys()),
        "study4_tasks": parse_novel_data_predictor(),
    }, ensure_ascii=False, indent=2))


def add_common(parser):
    parser.add_argument("--provider")
    parser.add_argument("--providers-config", default=str(MODEL_PROVIDERS_CONFIG))
    parser.add_argument("--base-url")
    parser.add_argument("--api-key")
    parser.add_argument("--model")
    parser.add_argument("--temperature", type=float)
    parser.add_argument("--max-tokens", type=int)
    parser.add_argument("--timeout", type=int)
    parser.add_argument("--max-concurrency", type=int)
    parser.add_argument("--sleep", type=float, default=0.0)
    parser.add_argument("--limit", type=int)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--out-dir")


def main():
    parser = argparse.ArgumentParser(description="Run method-flow reproductions of Lehr et al. 2024 with configurable OpenAI-compatible LLM APIs.")
    sub = parser.add_subparsers(dest="command", required=True)
    sub.add_parser("materials")
    p1 = sub.add_parser("study1")
    add_common(p1)
    p1.add_argument("--max-followups", type=int)
    p1.add_argument("--manual-count", action="store_true")
    p2 = sub.add_parser("study2")
    add_common(p2)
    p2.add_argument("--good", action="store_true")
    p2.add_argument("--single-turn", action="store_true")
    p3 = sub.add_parser("study3")
    add_common(p3)
    p4 = sub.add_parser("study4")
    add_common(p4)
    p4.add_argument("--repeats", type=int, default=5)
    p4.add_argument("--confirm-implicit-direction", action="store_true")
    args = parser.parse_args()
    if args.command == "materials":
        list_materials()
        return
    apply_model_provider_config(args)
    if args.timeout is None:
        args.timeout = 120
    client = LLMClient.from_args(args)
    {"study1": study1, "study2": study2, "study3": study3, "study4": study4}[args.command](args, client)


if __name__ == "__main__":
    main()
